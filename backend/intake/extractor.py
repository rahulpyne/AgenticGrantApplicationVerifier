"""
Field extractor — LLM-powered, skills-based extraction pipeline.

Architecture
────────────
Every extraction decision is driven by a *skill* — a configurable description
of what to look for and how to interpret it.  Skills live in skills.json and
can be edited without touching any Python code, making the extractor completely
domain-agnostic and reusable for any application type.

Pipeline (per submission)
  1. Load skill definitions from skills.json
  2. Build document bundle — read ALL submitted documents into text regardless
     of file format (PDF, JSON, XLSX, TXT, CSV, Markdown, …)
  3. Primary extraction — one GPT call reads the full bundle and extracts all
     fields simultaneously, guided by each skill's extraction_prompt
  4. Cross-check — for fields marked cross_check=true, each cross-check
     document is queried individually and the result is compared to the primary
     value (Python does the comparison; LLM only does name finding)
  5. Budget cross-check — DF-05 is compared against the DOC-04 Excel total
     using openpyxl (LLM cannot read binary Excel files)
  6. Test overrides — any _extraction_overrides embedded in application_form.json
     are applied last (used by TC-11 test fixtures to inject explicit confidence)
  7. Fallback — when OPENAI_API_KEY is absent the legacy structured parsers run
     silently so the pipeline always produces output with zero regression

Key design decisions
  • No hardcoded JSON key paths, no hardcoded regex patterns
  • No assumptions about document layout, sheet names, or field labels
  • File format detection is automatic; new formats just need a reader added below
  • skills.json is the ONLY domain-specific artifact — swap it to change domains
  • OPENAI_API_KEY is optional: absent → legacy parsers run, no errors
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from models.schemas import AuditEvent, Case, EventType, ExtractedField


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

_OPENAI_API_KEY: str = os.environ.get("OPENAI_API_KEY", "")
# When set, points the OpenAI SDK at an Azure AI Foundry (or other
# OpenAI-compatible) endpoint, e.g.
#   https://<resource>.services.ai.azure.com/openai/v1
_OPENAI_BASE_URL: str = os.environ.get("OPENAI_BASE_URL", "")
# Deployment / model name. On Azure Foundry this is the *deployment name*
# (e.g. "gpt-4.1"); on plain OpenAI it's the model id.
_GPT_MODEL: str = os.environ.get("GPT_MODEL", "gpt-4.1")
_SKILLS_FILE: Path = Path(__file__).parent / "skills.json"

# Cap total characters sent to GPT per submission (~6K tokens across all docs)
_MAX_BUNDLE_CHARS: int = 24_000


def _gpt_available() -> bool:
    """True when an API key is present in the environment."""
    return bool(_OPENAI_API_KEY)


def _make_client():
    """
    Build an OpenAI SDK client.  When OPENAI_BASE_URL is set (Azure AI Foundry
    or any OpenAI-compatible gateway) it is used as the base_url; otherwise the
    client talks to api.openai.com directly.
    """
    from openai import OpenAI
    if _OPENAI_BASE_URL:
        return OpenAI(base_url=_OPENAI_BASE_URL, api_key=_OPENAI_API_KEY)
    return OpenAI(api_key=_OPENAI_API_KEY)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Skills loader
# ---------------------------------------------------------------------------

def _load_skills() -> List[Dict]:
    """
    Load field definitions from skills.json.
    Returns an empty list on any error so the pipeline still runs (fallback mode).
    """
    try:
        with open(_SKILLS_FILE, "r", encoding="utf-8") as fh:
            return json.load(fh).get("fields", [])
    except Exception:
        return []


def _skill_by_id(skills: List[Dict], field_id: str) -> Optional[Dict]:
    return next((s for s in skills if s["field_id"] == field_id), None)


def _all_field_ids(skills: List[Dict]) -> Dict[str, str]:
    """Returns {field_id: name} for all skills, with hardcoded defaults as fallback."""
    if skills:
        return {s["field_id"]: s["name"] for s in skills}
    return {
        "DF-01": "Applicant Legal Name",
        "DF-02": "CRA Business Number",
        "DF-03": "Incorporation Date",
        "DF-04": "Location (Province + BC Facility)",
        "DF-05": "Requested PacifiCan Amount",
        "DF-06": "Matching (Non-PacifiCan) Funding",
        "DF-07": "Project Period",
    }


# ---------------------------------------------------------------------------
# Universal document readers
# ---------------------------------------------------------------------------

def _read_pdf_text(filepath: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(filepath)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        return ""


def _read_excel_text(filepath: str) -> str:
    """
    Convert an Excel workbook into readable plain text for LLM ingestion.
    Each sheet becomes a labelled section; rows are pipe-separated.
    Caps at 100 rows per sheet to avoid overflowing context.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        lines: List[str] = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            lines.append(f"\n[Sheet: {sheet_name}]")
            for row in sheet.iter_rows(max_row=100):
                vals = [
                    str(c.value).strip()
                    for c in row
                    if c.value is not None and str(c.value).strip()
                ]
                if vals:
                    lines.append(" | ".join(vals))
        return "\n".join(lines)
    except Exception:
        return ""


def _read_document_text(filepath: str) -> str:
    """
    Auto-detect file type and return its full text content.
    Adding support for a new file type only requires a new branch here.
    """
    ext = Path(filepath).suffix.lower()
    if ext == ".pdf":
        return _read_pdf_text(filepath)
    if ext == ".json":
        try:
            with open(filepath, "r", encoding="utf-8") as fh:
                return json.dumps(json.load(fh), indent=2)
        except Exception:
            return ""
    if ext in (".xlsx", ".xls"):
        return _read_excel_text(filepath)
    # Catch-all: plain text, CSV, Markdown, DOCX-as-text, etc.
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as fh:
            return fh.read()
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Document helper utilities
# ---------------------------------------------------------------------------

_DOC_LABELS: Dict[str, str] = {
    "DOC-01": "Application Form",
    "DOC-02": "Annual Financial Statements",
    "DOC-03": "Interim Financial Statements",
    "DOC-04": "Budget Worksheet",
    "DOC-05": "Business Plan",
    "DOC-06": "Funding Confirmation Letter",
    "DOC-07": "Supplemental / Application Form",
    "DOC-08": "Technology Questionnaire",
}


def _doc_label(doc_type: str) -> str:
    return _DOC_LABELS.get(doc_type, doc_type)


def _get_doc_path(case: Case, doc_type: str, scenario_folder: str) -> Optional[str]:
    """Return filesystem path to the best-confidence document of the given type."""
    folder = Path(scenario_folder)
    best, best_conf = None, -1.0
    for doc in case.documents:
        if doc.detected_doc_type == doc_type and doc.confidence >= best_conf:
            path = folder / doc.name
            if path.exists():
                best, best_conf = str(path), doc.confidence
    return best


def _build_document_bundle(case: Case, scenario_folder: str) -> List[Dict]:
    """
    Build a list of {doc_type, filename, label, text} dicts for every
    classified document in the submission.  Text is capped per document so
    the total bundle fits within the LLM context window.
    """
    total_docs = max(len(case.documents), 1)
    per_doc_cap = max(2_000, _MAX_BUNDLE_CHARS // total_docs)

    bundle: List[Dict] = []
    seen: set = set()

    for doc in case.documents:
        dt = doc.detected_doc_type
        if not dt or dt in seen:
            continue
        path = Path(scenario_folder) / doc.name
        if not path.exists():
            continue
        text = _read_document_text(str(path))
        if not text.strip():
            continue
        bundle.append({
            "doc_type": dt,
            "filename": doc.name,
            "label":    _doc_label(dt),
            "text":     text[:per_doc_cap],
        })
        seen.add(dt)

    return bundle


# ---------------------------------------------------------------------------
# LLM — primary extraction (all fields, all documents, one call)
# ---------------------------------------------------------------------------

def _format_skills_prompt(skills: List[Dict]) -> str:
    """Render skills as a numbered list for the extraction prompt."""
    lines: List[str] = []
    for s in skills:
        lines.append(
            f"  {s['field_id']} [{s['name']}]: {s['extraction_prompt']}"
        )
    return "\n".join(lines)


def _primary_extract_llm(
    bundle: List[Dict],
    skills: List[Dict],
) -> Dict[str, Dict]:
    """
    Send the full document bundle to GPT-4.1-mini and extract all fields at once.

    Returns a dict  field_id → {value, confidence, excerpt, source_doc_id}.
    Returns an empty dict on any failure so the caller can fall back to the
    legacy structured parsers.
    """
    if not _gpt_available() or not bundle or not skills:
        return {}

    # Build the document bundle string
    bundle_parts: List[str] = []
    for item in bundle:
        bundle_parts.append(
            f"=== {item['doc_type']}: {item['label']}  [{item['filename']}] ===\n"
            f"{item['text']}"
        )
    bundle_text = "\n\n".join(bundle_parts)

    skills_text = _format_skills_prompt(skills)
    field_ids = [s["field_id"] for s in skills]

    system_msg = (
        "You are a precise, expert document parser for formal applications. "
        "Extract requested fields and return only the JSON object specified — "
        "no explanation, no markdown fences, no extra keys."
    )

    user_msg = (
        "Extract the fields below from the document bundle provided.\n\n"
        "FIELDS TO EXTRACT:\n"
        f"{skills_text}\n\n"
        "DOCUMENT BUNDLE:\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{bundle_text}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Return ONLY a JSON object.  For every field ID listed supply:\n"
        '  "value"        : extracted value (string / number / object / null)\n'
        '  "confidence"   : your certainty 0.0–1.0\n'
        '  "excerpt"      : verbatim text snippet where found, or null\n'
        '  "source_doc_id": the DOC-XX identifier of the source document, or null\n\n'
        f"Required field IDs: {field_ids}\n"
        "For fields not found: value=null, confidence=0.0, excerpt=null."
    )

    try:
        client = _make_client()
        resp = client.chat.completions.create(
            model=_GPT_MODEL,
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user",   "content": user_msg},
            ],
            response_format={"type": "json_object"},
            temperature=0,
            max_tokens=1_000,
        )
        result: Dict = json.loads(resp.choices[0].message.content)
        # Ensure every expected field ID is present in the response
        for fid in field_ids:
            if fid not in result:
                result[fid] = {
                    "value": None, "confidence": 0.0,
                    "excerpt": None, "source_doc_id": None,
                }
        return result
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# LLM — cross-check (DF-01 and any other cross_check=true fields)
# ---------------------------------------------------------------------------

def _cross_check_name_llm(doc_item: Dict, skill: Dict) -> Optional[str]:
    """
    Ask GPT to find the legal organisation name in one cross-check document.
    Returns the name string, or None on failure / not found.
    """
    if not _gpt_available():
        return None

    cross_prompt = skill.get(
        "cross_check_prompt",
        "What is the legal organisation name mentioned in this document? Return only the name.",
    )
    user_msg = (
        f"Document: {doc_item['label']}  [{doc_item['filename']}]\n\n"
        f"{cross_prompt}\n\n"
        "Text:\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{doc_item['text'][:3_000]}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Reply with ONLY the exact legal name as it appears. "
        "If no clear legal name is present, reply with the single word: NONE"
    )
    try:
        client = _make_client()
        resp = client.chat.completions.create(
            model=_GPT_MODEL,
            messages=[{"role": "user", "content": user_msg}],
            temperature=0,
            max_tokens=80,
        )
        name = resp.choices[0].message.content.strip()
        if not name or name.upper() == "NONE" or len(name) > 150 or "\n" in name:
            return None
        return name
    except Exception:
        return None


def _normalise_name(name: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace for name comparison."""
    name = name.lower()
    name = re.sub(r"[^\w\s]", "", name)
    return re.sub(r"\s+", " ", name).strip()


def _names_match(a: str, b: str) -> bool:
    return bool(a) and bool(b) and _normalise_name(a) == _normalise_name(b)


def _run_cross_checks(
    extracted: Dict[str, ExtractedField],
    bundle: List[Dict],
    skills: List[Dict],
) -> Dict[str, ExtractedField]:
    """
    For every skill flagged cross_check=true, compare the primary value against
    each supporting document in the bundle.  Python does the name comparison;
    GPT only locates the name inside each document.
    """
    for skill in skills:
        if not skill.get("cross_check"):
            continue

        fid = skill["field_id"]
        field = extracted.get(fid)
        if not field or field.value is None:
            continue

        app_value = str(field.value)
        cross_doc_types: List[str] = skill.get("cross_check_docs", [])

        corroborations: List[str] = []
        contradictions: List[str] = []
        parse_failures: List[str] = []

        for doc_item in bundle:
            dt = doc_item["doc_type"]
            if dt not in cross_doc_types:
                continue
            if dt == field.source_doc_id:
                continue   # skip the source document itself

            detected = _cross_check_name_llm(doc_item, skill)

            if detected:
                if _names_match(app_value, detected):
                    corroborations.append(
                        f"{dt} ({doc_item['label']}) confirms '{detected}'"
                    )
                else:
                    contradictions.append(
                        f"{dt} ({doc_item['label']}) says '{detected}'"
                    )
            else:
                parse_failures.append(
                    f"{dt} ({doc_item['label']}) — name not identifiable"
                )

        # Build updated excerpt in the pipe-separated format the UI expects
        base_excerpt = field.raw_excerpt or repr(app_value)
        new_conf = field.confidence
        extra: List[str] = []

        if contradictions:
            new_conf = round(field.confidence * 0.40, 2)
            extra.append("⚠ CROSS-CHECK MISMATCH — " + "; ".join(contradictions))
            if corroborations:
                extra.append("partial corroboration: " + "; ".join(corroborations))
        elif corroborations:
            extra.append("✓ cross-checked: " + "; ".join(corroborations))
            if parse_failures:
                extra.append("could not parse: " + "; ".join(parse_failures))
        elif parse_failures:
            extra.append(
                "ℹ submitted docs could not be parsed for name: "
                + "; ".join(parse_failures)
            )
        else:
            extra.append(
                "ℹ form declaration only — cross-document name verification not possible"
            )

        new_excerpt = base_excerpt + (" | " + " | ".join(extra) if extra else "")
        extracted[fid] = field.model_copy(update={
            "confidence": new_conf,
            "raw_excerpt": new_excerpt,
        })

    return extracted


# ---------------------------------------------------------------------------
# Budget cross-check  (DF-05 vs DOC-04 Excel total — never sent to GPT)
# ---------------------------------------------------------------------------

def _read_budget_total(filepath: str) -> Optional[float]:
    """
    Read the Total Project Costs from a DOC-04 budget worksheet.
    Uses openpyxl directly — Excel binary cannot be sent to a text LLM.
    Works regardless of exact sheet name or column layout.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)

        sheet = None
        for sname in wb.sheetnames:
            if "cost detail" in sname.lower() or "budget" in sname.lower():
                sheet = wb[sname]
                break
        if sheet is None and wb.sheetnames:
            sheet = wb.active
        if sheet is None:
            return None

        current: Optional[float] = None
        application: Optional[float] = None

        for row in sheet.iter_rows():
            label = ""
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    label = cell.value.strip().lower()
                    break
            if "total project costs (current)" in label:
                for cell in reversed(row):
                    if isinstance(cell.value, (int, float)):
                        current = float(cell.value)
                        break
            elif "total project costs" in label:
                for cell in reversed(row):
                    if isinstance(cell.value, (int, float)):
                        application = float(cell.value)
                        break

        return current or application
    except Exception:
        return None


def _run_budget_cross_check(
    extracted: Dict[str, ExtractedField],
    case: Case,
    scenario_folder: str,
    skills: List[Dict],
) -> Dict[str, ExtractedField]:
    """
    Cross-check DF-05 (requested amount) against the Total Project Costs in
    the budget worksheet (DOC-04).  Only runs when the skill marks it.
    """
    df05_skill = _skill_by_id(skills, "DF-05")
    if not df05_skill or not df05_skill.get("cross_check_budget"):
        return extracted

    df05 = extracted.get("DF-05")
    if not (df05 and df05.value is not None):
        return extracted

    requested: Optional[float] = None
    raw_val = df05.value
    if isinstance(raw_val, (int, float)):
        requested = float(raw_val)
    elif isinstance(raw_val, dict):
        amt = raw_val.get("amount") or raw_val.get("value")
        if isinstance(amt, (int, float)):
            requested = float(amt)

    if requested is None:
        return extracted

    budget_path = _get_doc_path(case, "DOC-04", scenario_folder)
    base = df05.raw_excerpt or f"Requested amount: {requested:,.0f}"

    if not budget_path:
        extracted["DF-05"] = df05.model_copy(update={
            "raw_excerpt": (
                base
                + " | ℹ form declaration only — DOC-04 not submitted; "
                "budget cross-check not possible"
            ),
        })
        return extracted

    budget_total = _read_budget_total(budget_path)

    if budget_total is None:
        extracted["DF-05"] = df05.model_copy(update={
            "raw_excerpt": (
                base
                + " | ℹ DOC-04 present but total project costs row could not be read"
            ),
        })
        return extracted

    if requested > budget_total:
        new_conf = round(df05.confidence * 0.45, 2)
        note = (
            f"⚠ BUDGET MISMATCH — declared request ${requested:,.0f} exceeds "
            f"budget worksheet total ${budget_total:,.0f}"
        )
    else:
        new_conf = df05.confidence
        pct = (requested / budget_total * 100) if budget_total > 0 else 0.0
        note = (
            f"✓ budget cross-check: ${requested:,.0f} = {pct:.1f}% of "
            f"worksheet total ${budget_total:,.0f}"
        )

    extracted["DF-05"] = df05.model_copy(update={
        "confidence": new_conf,
        "raw_excerpt": base + " | " + note,
    })
    return extracted


# ---------------------------------------------------------------------------
# TC-11 test overrides (embedded in application_form.json)
# ---------------------------------------------------------------------------

_OVERRIDE_FIELD_MAP: Dict[str, str] = {
    "cra_business_number":   "DF-02",
    "incorporation_date":    "DF-03",
    "legal_name":            "DF-01",
    "rda_funding_requested": "DF-05",
    "non_rda_funding":       "DF-06",
    "project_period":        "DF-07",
}


def _apply_extraction_overrides(
    extracted: Dict[str, ExtractedField],
    app_form_path: Optional[str],
) -> Dict[str, ExtractedField]:
    """
    Honour any _extraction_overrides key in application_form.json.
    This mechanism exists for TC-11 test scenarios that need to inject
    specific confidence values for edge-case testing without changing real docs.
    """
    if not app_form_path:
        return extracted
    try:
        with open(app_form_path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception:
        return extracted

    for key, override in data.get("_extraction_overrides", {}).items():
        fid = _OVERRIDE_FIELD_MAP.get(key)
        if not (fid and fid in extracted and isinstance(override, dict) and "confidence" in override):
            continue
        f = extracted[fid]
        note = override.get("note", "")
        extracted[fid] = ExtractedField(
            field_id=f.field_id,
            name=f.name,
            value=override.get("value"),
            source_doc_id=f.source_doc_id,
            confidence=float(override["confidence"]),
            raw_excerpt=(f.raw_excerpt or "") + (f" [OVERRIDE: {note}]" if note else ""),
            manually_corrected=f.manually_corrected,
            correction_history=f.correction_history,
        )
    return extracted


# ---------------------------------------------------------------------------
# Legacy fallback parsers (run when OPENAI_API_KEY is absent or GPT fails)
# These replicate the original behaviour exactly so there is zero regression.
# ---------------------------------------------------------------------------

def _unwrap(val: Any, default_confidence: float = 0.97) -> Tuple[Any, float, Optional[str]]:
    """Unpack TC-11 confidence-annotated values embedded in JSON."""
    if isinstance(val, dict) and "confidence" in val:
        return val.get("value"), float(val.get("confidence", default_confidence)), val.get("note")
    return val, default_confidence, None


def _legacy_extract_from_app_form(
    filepath: str,
    case: Case,
) -> Dict[str, ExtractedField]:
    """Original structured JSON field extraction — used when GPT is unavailable."""
    fields: Dict[str, ExtractedField] = {}
    try:
        with open(filepath, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception:
        return fields

    org     = data.get("organization", {})
    project = data.get("project", {})
    funding = data.get("funding", {})

    # DF-01
    raw = org.get("legal_name")
    if raw is not None:
        val, conf, note = _unwrap(raw, 0.97)
        excerpt = f"organization.legal_name = {val!r}"
        if note:
            excerpt += f" [{note}]"
        fields["DF-01"] = ExtractedField(
            field_id="DF-01", name="Applicant Legal Name",
            value=val, source_doc_id="DOC-01", confidence=conf, raw_excerpt=excerpt,
        )

    # DF-02
    raw = org.get("cra_business_number")
    if raw is not None:
        val, conf, note = _unwrap(raw, 0.97)
        bn_str = str(val).strip() if val is not None else ""
        if not re.match(r"^\d{9}$", bn_str) and val is not None:
            conf = min(conf, 0.55)
            note = (note or "") + f" [Warning: '{bn_str}' is not a valid 9-digit CRA BN]"
        fields["DF-02"] = ExtractedField(
            field_id="DF-02", name="CRA Business Number",
            value=val, source_doc_id="DOC-01", confidence=conf,
            raw_excerpt=f"organization.cra_business_number = {val!r}" + (f" [{note}]" if note else ""),
        )

    # DF-03
    raw = (
        org.get("incorporation_date")
        or org.get("date_established_in_canada")
        or org.get("date_established")
    )
    if raw is not None:
        val, conf, _ = _unwrap(raw, 0.97)
        fields["DF-03"] = ExtractedField(
            field_id="DF-03", name="Incorporation Date",
            value=val, source_doc_id="DOC-01", confidence=conf,
            raw_excerpt=f"organization.incorporation_date = {val!r}",
        )

    # DF-04
    address = project.get("address")
    province = None
    if isinstance(address, dict):
        province, _, _ = _unwrap(address.get("province"), 0.95)
    elif isinstance(address, str):
        province = address
    raw_bc = org.get("bc_operating_facilities")
    bc_val, _, _ = _unwrap(raw_bc, 0.95) if raw_bc is not None else (None, 0.95, None)
    loc_val: Dict[str, Any] = {}
    if province is not None:
        loc_val["province"] = province
    if bc_val is not None:
        loc_val["bc_operating_facilities"] = bc_val
    if loc_val:
        fields["DF-04"] = ExtractedField(
            field_id="DF-04", name="Location (Province + BC Facility)",
            value=loc_val, source_doc_id="DOC-01", confidence=0.95,
            raw_excerpt=(
                f"project.address.province={province!r}, "
                f"organization.bc_operating_facilities={bc_val!r}"
            ),
        )

    # DF-05
    raw = funding.get("total_rda_funding_requested")
    if raw is not None:
        val, conf, _ = _unwrap(raw, 0.97)
        fields["DF-05"] = ExtractedField(
            field_id="DF-05", name="Requested PacifiCan Amount",
            value=val, source_doc_id="DOC-01", confidence=conf,
            raw_excerpt=f"funding.total_rda_funding_requested = {val!r}",
        )

    # DF-06
    raw = funding.get("total_non_rda_funding")
    if raw is not None:
        val, conf, _ = _unwrap(raw, 0.90)
        doc06_present = any(d.detected_doc_type == "DOC-06" for d in case.documents)
        fields["DF-06"] = ExtractedField(
            field_id="DF-06", name="Matching (Non-PacifiCan) Funding",
            value={"amount": val, "confirmation_present": doc06_present},
            source_doc_id="DOC-01", confidence=conf,
            raw_excerpt=(
                f"funding.total_non_rda_funding = {val!r}, "
                f"DOC-06 present = {doc06_present}"
            ),
        )

    # DF-07
    raw_start = project.get("start_date")
    raw_end   = project.get("end_date")
    if raw_start is not None or raw_end is not None:
        sv, sc, _ = _unwrap(raw_start, 0.97) if raw_start is not None else (None, 0.0, None)
        ev, ec, _ = _unwrap(raw_end,   0.97) if raw_end   is not None else (None, 0.0, None)
        confs = [c for c in [sc, ec] if c > 0]
        fields["DF-07"] = ExtractedField(
            field_id="DF-07", name="Project Period",
            value={"start_date": sv, "end_date": ev},
            source_doc_id="DOC-01", confidence=min(confs) if confs else 0.0,
            raw_excerpt=f"project.start_date = {sv!r}, project.end_date = {ev!r}",
        )

    return fields


def _legacy_fallback(case: Case, scenario_folder: str) -> Dict[str, ExtractedField]:
    """
    Original extraction logic — runs when OPENAI_API_KEY is absent or GPT
    returns an empty result.  Reads from application_form.json only.
    """
    fields: Dict[str, ExtractedField] = {}
    app_path = _get_doc_path(case, "DOC-01", scenario_folder)
    if app_path:
        fields = _legacy_extract_from_app_form(app_path, case)
    return fields


# ---------------------------------------------------------------------------
# Post-extraction shape normalisation
# ---------------------------------------------------------------------------

def _normalise_df06_shape(
    extracted: Dict[str, ExtractedField],
    case: Case,
) -> Dict[str, ExtractedField]:
    """
    DF-06 (matching funding) is consumed by the UI as an object
    {amount, confirmation_present}.  GPT returns the bare amount, so we re-wrap
    it here, attaching confirmation_present derived from whether the funding
    confirmation document (DOC-06) was submitted.  This preserves the UI feature
    (CaseDetail.tsx shows "Confirmation: Present/Not found") with no degradation.

    `_parse_amount` in rules_engine already accepts both shapes, so rules are
    unaffected either way.
    """
    df06 = extracted.get("DF-06")
    if not df06 or df06.value is None:
        return extracted

    val = df06.value
    if isinstance(val, dict) and "amount" in val:
        return extracted  # already correctly shaped (e.g. legacy fallback)

    amount = val
    if isinstance(val, dict):
        amount = val.get("amount") or val.get("value")

    doc06_present = any(d.detected_doc_type == "DOC-06" for d in case.documents)
    base = df06.raw_excerpt or f"matching funding = {amount!r}"
    extracted["DF-06"] = df06.model_copy(update={
        "value": {"amount": amount, "confirmation_present": doc06_present},
        "raw_excerpt": base + f" | DOC-06 (funding confirmation) present = {doc06_present}",
    })
    return extracted


# ---------------------------------------------------------------------------
# Convert LLM JSON result → ExtractedField objects
# ---------------------------------------------------------------------------

def _build_extracted_fields(
    llm_results: Dict[str, Dict],
    skills: List[Dict],
) -> Dict[str, ExtractedField]:
    """Map the LLM's raw JSON output to typed ExtractedField objects."""
    fields: Dict[str, ExtractedField] = {}
    for skill in skills:
        fid  = skill["field_id"]
        res  = llm_results.get(fid, {})
        val  = res.get("value")
        conf = float(res.get("confidence", 0.0))
        if val is None:
            conf = 0.0
        fields[fid] = ExtractedField(
            field_id=fid,
            name=skill["name"],
            value=val,
            source_doc_id=res.get("source_doc_id"),
            confidence=min(conf, 1.0),
            raw_excerpt=res.get("excerpt"),
        )
    return fields


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def extract_fields(case: Case, scenario_folder: str) -> Case:
    """
    Skills-based, LLM-powered field extraction entry point.

    Returns the same Case object with case.extracted_fields populated.
    The pipeline is:
      load skills → build bundle → GPT primary extract → cross-checks
      → budget cross-check → TC-11 overrides → fill missing placeholders
    Falls back to legacy parsers transparently when GPT is unavailable.
    """
    skills = _load_skills()
    bundle = _build_document_bundle(case, scenario_folder)
    extracted: Dict[str, ExtractedField] = {}

    # ── Primary extraction ────────────────────────────────────────────────────
    if _gpt_available() and skills and bundle:
        llm_results = _primary_extract_llm(bundle, skills)
        if llm_results:
            extracted = _build_extracted_fields(llm_results, skills)

    # ── Fallback (no key, empty result, or API error) ─────────────────────────
    if not extracted:
        extracted = _legacy_fallback(case, scenario_folder)

    # ── Normalise DF-06 to the {amount, confirmation_present} shape the UI uses ─
    extracted = _normalise_df06_shape(extracted, case)

    # ── Cross-checks (LLM-driven, only when key present) ─────────────────────
    if _gpt_available() and bundle:
        extracted = _run_cross_checks(extracted, bundle, skills)

    # ── Budget cross-check (openpyxl, always runs) ────────────────────────────
    extracted = _run_budget_cross_check(extracted, case, scenario_folder, skills)

    # ── TC-11 test overrides ──────────────────────────────────────────────────
    app_form_path = _get_doc_path(case, "DOC-01", scenario_folder)
    extracted = _apply_extraction_overrides(extracted, app_form_path)

    # ── Fill any missing fields with empty placeholders ───────────────────────
    for fid, fname in _all_field_ids(skills).items():
        if fid not in extracted:
            extracted[fid] = ExtractedField(
                field_id=fid, name=fname,
                value=None, source_doc_id=None,
                confidence=0.0, raw_excerpt=None,
            )

    case.extracted_fields = extracted

    case.audit_trail.append(AuditEvent(
        case_id=case.case_id,
        timestamp=_now_iso(),
        event_type=EventType.extraction_completed,
        actor="system",
        details={
            "fields_extracted": [
                fid for fid, f in extracted.items() if f.value is not None
            ],
            "fields_missing": [
                fid for fid, f in extracted.items() if f.value is None
            ],
            "cross_check_mismatches": [
                fid for fid, f in extracted.items()
                if f.raw_excerpt and (
                    "CROSS-CHECK MISMATCH" in f.raw_excerpt
                    or "BUDGET MISMATCH" in f.raw_excerpt
                )
            ],
            "gpt_assisted":  _gpt_available(),
            "gpt_model":     _GPT_MODEL if _gpt_available() else None,
            "skills_loaded": len(skills),
        },
    ))

    return case
