"""
Document classifier — maps DocumentRecord objects to DOC-01 through DOC-08.
Uses filename heuristics first, then content signatures where possible.
"""
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Tuple

from models.schemas import AuditEvent, Case, ChecklistItem, DocumentRecord, DocumentStatus, EventType

# ---------------------------------------------------------------------------
# Checklist catalogue
# ---------------------------------------------------------------------------
CHECKLIST_CATALOGUE = {
    "DOC-01": "Application Form (JSON)",
    "DOC-02": "Annual Financial Statements",
    "DOC-03": "Interim Financial Statements",
    "DOC-04": "Budget Worksheet (XLSX)",
    "DOC-05": "Business Plan / Pitch Deck",
    "DOC-06": "Funding Confirmation Letter",
    "DOC-07": "RDII Mandatory Supplemental Form",
    "DOC-08": "RDII Technology Questionnaire",
}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _stem(name: str) -> str:
    """Return the lowercase filename stem with separators normalised to spaces."""
    stem = Path(name).stem.lower()
    return re.sub(r"[_\-\s]+", " ", stem).strip()


def _classify_one(
    record: DocumentRecord,
    scenario_folder: str,
    already_assigned: dict[str, list[DocumentRecord]],
) -> Tuple[Optional[str], float, str, Optional[str]]:
    """
    Return (detected_doc_type, confidence, matched_on, notes).
    already_assigned maps doc_type → list of DocumentRecord already mapped to it.
    """
    name_lower = record.name.lower()
    stem = _stem(record.name)
    ext = record.extension.lower()
    file_path = Path(scenario_folder) / record.name
    notes: Optional[str] = None

    # ------------------------------------------------------------------
    # DOC-01: Application form
    # ------------------------------------------------------------------
    if ext == ".json":
        # Try content check
        try:
            with open(file_path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if "case_id" in data:
                return "DOC-01", 0.98, "content", None
        except Exception:
            pass
        if "application" in stem or "application form" in stem:
            return "DOC-01", 0.95, "filename", None
        return "DOC-01", 0.80, "filename", None

    # ------------------------------------------------------------------
    # DOC-03: Interim financials (must check BEFORE DOC-02)
    # ------------------------------------------------------------------
    if "interim" in stem:
        return "DOC-03", 0.90, "filename", None

    # ------------------------------------------------------------------
    # DOC-08: Technology questionnaire
    # ------------------------------------------------------------------
    if (
        "tech q" in stem
        or "technology questionnaire" in stem
        or "techq" in stem
        or ("tech" in stem and ("quest" in stem or "questionnaire" in stem))
    ):
        return "DOC-08", 0.90, "filename", None

    # ------------------------------------------------------------------
    # DOC-02: Annual financial statements
    # ------------------------------------------------------------------
    is_financial = (
        "financial statement" in stem
        or "financial statements" in stem
        or ("financials" in stem and "interim" not in stem)
    )
    if is_financial:
        # TC-12 duplicate/ambiguous: if a file named "financials_YYYY.pdf" is found
        # AND there is already a DOC-02 mapped
        ambiguous_year = re.search(r"\bfinancials[ _]20\d{2}\b", name_lower)
        if ambiguous_year and "DOC-02" in already_assigned and already_assigned["DOC-02"]:
            notes = (
                "Uncertain — possible duplicate or second year of annual statements. "
                "Manual disambiguation required."
            )
            return "uncertain_financial", 0.55, "filename", notes

        # Straightforward financial statement
        conf = 0.85
        if re.search(r"\b(annual|year.?end|20\d{2})\b", stem):
            conf = 0.85
        elif "financials" in stem and not re.search(r"\b20\d{2}\b", stem):
            conf = 0.75
        return "DOC-02", conf, "filename", None

    # ------------------------------------------------------------------
    # DOC-04: Budget worksheet
    # ------------------------------------------------------------------
    if "budget" in stem:
        if ext == ".xlsx":
            # Try reading tab names for content confirmation
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
                if "Cost Detail" in wb.sheetnames:
                    return "DOC-04", 0.95, "content", None
                return "DOC-04", 0.90, "filename", None
            except Exception:
                pass
        return "DOC-04", 0.90, "filename", None

    if ext in {".xlsx", ".xls"}:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            if "Cost Detail" in wb.sheetnames:
                return "DOC-04", 0.95, "content", None
        except Exception:
            pass
        return "DOC-04", 0.90, "filename", None

    # ------------------------------------------------------------------
    # DOC-05: Business plan / pitch deck
    # ------------------------------------------------------------------
    if "business plan" in stem or "business_plan" in stem:
        return "DOC-05", 0.90, "filename", None
    if "pitch" in stem:
        return "DOC-05", 0.90, "filename", None
    if "business" in stem and "plan" in stem:
        return "DOC-05", 0.80, "filename", None

    # ------------------------------------------------------------------
    # DOC-06: Funding confirmation
    # ------------------------------------------------------------------
    if "funding confirmation" in stem or "funding_confirmation" in stem:
        return "DOC-06", 0.85, "filename", None
    if "confirmation" in stem:
        return "DOC-06", 0.85, "filename", None

    # ------------------------------------------------------------------
    # DOC-07: Supplemental form
    # ------------------------------------------------------------------
    if "supplemental" in stem:
        return "DOC-07", 0.90, "filename", None

    # Could not classify
    return None, 0.0, "filename", None


def _build_checklist(
    classified: list[DocumentRecord],
    tech_commercialization: bool,
) -> list[ChecklistItem]:
    """
    Build the 7 or 8 checklist items based on classification results.
    """
    checklist: list[ChecklistItem] = []
    required_docs = list(CHECKLIST_CATALOGUE.keys())
    if not tech_commercialization:
        required_docs = [d for d in required_docs if d != "DOC-08"]

    # Group records by doc_type
    by_type: dict[str, list[DocumentRecord]] = {}
    for doc in classified:
        dt = doc.detected_doc_type
        if dt:
            by_type.setdefault(dt, []).append(doc)

    for doc_id in required_docs:
        category = CHECKLIST_CATALOGUE[doc_id]
        matched = by_type.get(doc_id, [])

        if matched:
            # Use the highest-confidence match
            best = max(matched, key=lambda d: d.confidence)
            if best.confidence >= 0.70:
                status = DocumentStatus.present
            else:
                status = DocumentStatus.uncertain
            checklist.append(ChecklistItem(
                doc_id=doc_id,
                category=category,
                status=status,
                matched_files=[r.name for r in matched],
                confidence=best.confidence,
                notes=best.notes,
            ))
        else:
            # Check for uncertain_financial that might be DOC-03
            if doc_id == "DOC-03":
                uncertain_fin = by_type.get("uncertain_financial", [])
                if uncertain_fin:
                    best = uncertain_fin[0]
                    checklist.append(ChecklistItem(
                        doc_id=doc_id,
                        category=category,
                        status=DocumentStatus.uncertain,
                        matched_files=[r.name for r in uncertain_fin],
                        confidence=best.confidence,
                        notes=best.notes,
                    ))
                    continue

            checklist.append(ChecklistItem(
                doc_id=doc_id,
                category=category,
                status=DocumentStatus.missing,
                matched_files=[],
                confidence=0.0,
            ))

    # DOC-08 if tech_commercialization is False — add as not_applicable
    if not tech_commercialization:
        checklist.append(ChecklistItem(
            doc_id="DOC-08",
            category=CHECKLIST_CATALOGUE["DOC-08"],
            status=DocumentStatus.not_applicable,
            matched_files=[],
            confidence=0.0,
        ))

    return checklist


def _classify_documents_legacy(case: Case, scenario_folder: str) -> Case:
    """
    DEACTIVATED by default — retained as the no-key fallback.

    Deterministic filename/content classification. Runs only when no model key
    is configured (see classify_documents dispatcher below).
    """
    already_assigned: dict[str, list[DocumentRecord]] = {}
    classified: list[DocumentRecord] = []

    for record in case.documents:
        doc_type, confidence, matched_on, notes = _classify_one(
            record, scenario_folder, already_assigned
        )
        updated = record.model_copy(update={
            "detected_doc_type": doc_type,
            "confidence": confidence,
            "matched_on": matched_on,
            "parse_status": "parsed" if doc_type else "unclassified",
            "notes": notes,
        })
        classified.append(updated)
        if doc_type:
            already_assigned.setdefault(doc_type, []).append(updated)

    case.documents = classified

    # Determine if technology commercialization applies
    tech_comm = _is_tech_commercialization(case, scenario_folder)

    case.checklist = _build_checklist(classified, tech_comm)

    audit = AuditEvent(
        case_id=case.case_id,
        timestamp=_now_iso(),
        event_type=EventType.classification_completed,
        actor="system",
        details={
            "engine": "rules-legacy",
            "classified_count": sum(1 for d in classified if d.detected_doc_type),
            "unclassified_count": sum(1 for d in classified if not d.detected_doc_type),
            "tech_commercialization": tech_comm,
        },
    )
    case.audit_trail.append(audit)
    return case


def _is_tech_commercialization(case: Case, scenario_folder: str) -> bool:
    """Read application_form.json to determine technology_commercialization flag."""
    app_form = Path(scenario_folder) / "application_form.json"
    if app_form.exists():
        try:
            with open(app_form, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            project = data.get("project", {})
            return bool(project.get("technology_commercialization", False))
        except Exception:
            pass
    return False


# ===========================================================================
# MODEL / SKILL-BASED classification (default path when a key is configured)
# ===========================================================================

_DOCS_SKILL_FILE = Path(__file__).parent / "documents.json"


def _load_doc_skills() -> dict:
    try:
        with open(_DOCS_SKILL_FILE, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def _classify_documents_llm(case: Case, scenario_folder: str) -> Case:
    """
    Model/skill-based classification + completeness.

    The model reads every uploaded file (filename + content) and, guided by
    documents.json, assigns each to a DOC-type and produces the completeness
    checklist (present / uncertain / missing / not_applicable) directly — no
    deterministic filename rules.  Raises on failure so the dispatcher can
    fall back to the dormant legacy classifier.
    """
    from intake import llm
    from intake.extractor import _read_document_text  # universal reader

    skills = _load_doc_skills()
    doc_types = skills.get("doc_types", [])

    folder = Path(scenario_folder)
    files_payload = []
    per_file_cap = 3_000
    for record in case.documents:
        path = folder / record.name
        text = _read_document_text(str(path)) if path.exists() else ""
        files_payload.append({
            "filename": record.name,
            "extension": record.extension,
            "content_preview": text[:per_file_cap],
        })

    system_msg = (
        "You are a precise intake classifier for formal application packages. "
        "Assign each uploaded file to the correct document type and assess "
        "package completeness. Return only the specified JSON — no prose."
    )
    user_msg = (
        "DOCUMENT TYPE DEFINITIONS (the only domain knowledge):\n"
        f"{json.dumps(doc_types, indent=2)}\n\n"
        "COMPLETENESS RULES:\n"
        f"{json.dumps(skills.get('completeness_rules', {}), indent=2)}\n\n"
        "UPLOADED FILES:\n"
        f"{json.dumps(files_payload, indent=2)}\n\n"
        "Return ONLY this JSON object:\n"
        "{\n"
        '  "tech_commercialization": <bool — from the application form project.technology_commercialization>,\n'
        '  "files": [ {"filename": <str>, "doc_id": <"DOC-0X" or null>, "confidence": <0..1>, "notes": <str or null>} ],\n'
        '  "checklist": [ {"doc_id": <"DOC-0X">, "category": <str>, "status": <"present"|"uncertain"|"missing"|"not_applicable">, "matched_files": [<str>], "confidence": <0..1>, "notes": <str or null>} ]\n'
        "}\n"
        "Include a checklist entry for every defined doc type (DOC-01..DOC-08). "
        "DOC-08 must be 'not_applicable' when tech_commercialization is false."
    )

    result = llm.chat_json(system_msg, user_msg, max_tokens=2000)
    if not result or "files" not in result or "checklist" not in result:
        raise RuntimeError("LLM classification returned no usable result")

    by_name = {f.get("filename"): f for f in result.get("files", [])}
    classified: list[DocumentRecord] = []
    for record in case.documents:
        info = by_name.get(record.name, {})
        doc_id = info.get("doc_id")
        conf = float(info.get("confidence", 0.0) or 0.0)
        classified.append(record.model_copy(update={
            "detected_doc_type": doc_id,
            "confidence": conf,
            "matched_on": "model",
            "parse_status": "parsed" if doc_id else "unclassified",
            "notes": info.get("notes"),
        }))
    case.documents = classified

    tech_comm = bool(result.get("tech_commercialization", False))

    cat_labels = {d["doc_id"]: d["category"] for d in doc_types}
    status_map = {
        "present": DocumentStatus.present,
        "uncertain": DocumentStatus.uncertain,
        "missing": DocumentStatus.missing,
        "not_applicable": DocumentStatus.not_applicable,
    }
    checklist: list[ChecklistItem] = []
    seen_ids: set[str] = set()
    for item in result.get("checklist", []):
        doc_id = item.get("doc_id")
        if not doc_id or doc_id in seen_ids:
            continue
        seen_ids.add(doc_id)
        checklist.append(ChecklistItem(
            doc_id=doc_id,
            category=item.get("category") or cat_labels.get(doc_id, doc_id),
            status=status_map.get(str(item.get("status")), DocumentStatus.missing),
            matched_files=item.get("matched_files", []) or [],
            confidence=float(item.get("confidence", 0.0) or 0.0),
            notes=item.get("notes"),
        ))
    for d in doc_types:
        if d["doc_id"] not in seen_ids:
            default_status = (
                DocumentStatus.not_applicable
                if d["doc_id"] == "DOC-08" and not tech_comm
                else DocumentStatus.missing
            )
            checklist.append(ChecklistItem(
                doc_id=d["doc_id"], category=d["category"],
                status=default_status, matched_files=[], confidence=0.0,
            ))
    case.checklist = checklist

    case.audit_trail.append(AuditEvent(
        case_id=case.case_id,
        timestamp=_now_iso(),
        event_type=EventType.classification_completed,
        actor="system",
        details={
            "engine": "model",
            "classified_count": sum(1 for d in classified if d.detected_doc_type),
            "unclassified_count": sum(1 for d in classified if not d.detected_doc_type),
            "tech_commercialization": tech_comm,
        },
    ))
    return case


def classify_documents(case: Case, scenario_folder: str) -> Case:
    """
    Dispatcher: model/skill-based classification when a key is configured,
    otherwise the (dormant) legacy deterministic classifier as a safety net.
    """
    from intake import llm
    if llm.gpt_available():
        try:
            return _classify_documents_llm(case, scenario_folder)
        except Exception:
            return _classify_documents_legacy(case, scenario_folder)
    return _classify_documents_legacy(case, scenario_folder)
