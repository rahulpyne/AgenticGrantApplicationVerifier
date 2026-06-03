"""
Rules engine — evaluates all R-001 through R-012 rules.
"""
from __future__ import annotations

import json
import os
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from models.schemas import (
    AuditEvent,
    Basket,
    Case,
    DocumentStatus,
    EligibilityFlag,
    EventType,
    ExtractedField,
    Finding,
    Severity,
)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalise_name(name: Optional[str]) -> str:
    """Lowercase, remove punctuation, collapse whitespace."""
    if not name:
        return ""
    name = name.lower()
    name = re.sub(r"[^\w\s]", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def _get_field(fields: Dict[str, ExtractedField], field_id: str) -> Optional[ExtractedField]:
    return fields.get(field_id)


def _parse_amount(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict):
        # DF-06 wraps amount in a dict
        v = value.get("amount") or value.get("value")
        if v is not None:
            return _parse_amount(v)
        return None
    s = str(value).replace(",", "").replace("$", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        from dateutil.parser import parse as dateutil_parse
        try:
            return dateutil_parse(value).date()
        except Exception:
            return None
    return None


def _read_pdf_text(filepath: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(filepath)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        return ""


def _get_file_path(case: Case, doc_type: str, submission_folder: str) -> Optional[str]:
    """Return filesystem path to the first document matching doc_type."""
    folder = Path(submission_folder)
    for doc in case.documents:
        if doc.detected_doc_type == doc_type:
            p = folder / doc.name
            if p.exists():
                return str(p)
    return None


def _unwrap(val: Any) -> Any:
    """Unwrap TC-11 style {value: ..., confidence: ...} dicts."""
    if isinstance(val, dict) and "value" in val:
        return val["value"]
    return val


def _load_app_form(case: Case, submission_folder: str) -> Optional[Dict]:
    path = _get_file_path(case, "DOC-01", submission_folder)
    if path and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            pass
    return None


# ---------------------------------------------------------------------------
# Budget worksheet helpers (R-003)
# ---------------------------------------------------------------------------

def _read_budget_totals(filepath: str) -> Tuple[Optional[float], Optional[float]]:
    """
    Read Total Project Costs (Current) and (Application) from the Cost Detail tab.
    Returns (current_total, application_total).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = None
        for name in wb.sheetnames:
            if "cost detail" in name.lower():
                sheet = wb[name]
                break
        if sheet is None and wb.sheetnames:
            sheet = wb.active
        if sheet is None:
            return None, None

        current_total: Optional[float] = None
        application_total: Optional[float] = None

        for row in sheet.iter_rows():
            row_label = ""
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    row_label = cell.value.strip()
                    break
            if not row_label:
                continue

            if "total project costs (current)" in row_label.lower():
                for cell in reversed(row):
                    if isinstance(cell.value, (int, float)):
                        current_total = float(cell.value)
                        break
            elif "total project costs (application)" in row_label.lower():
                for cell in reversed(row):
                    if isinstance(cell.value, (int, float)):
                        application_total = float(cell.value)
                        break

        return current_total, application_total
    except Exception:
        return None, None


def _read_total_project_costs_from_budget(filepath: str) -> Optional[float]:
    current, application = _read_budget_totals(filepath)
    return current or application


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def _run_rules_legacy(case: Case, submission_folder: str) -> Case:
    """
    DEACTIVATED by default — retained as the no-key fallback.

    Deterministic evaluation of R-001 through R-012 plus ER-01..ER-09.
    Runs only when no model key is configured (see run_rules dispatcher).
    """
    findings: List[Finding] = list(case.findings)  # preserve existing (e.g. R-002 from extractor)
    flags: List[EligibilityFlag] = []

    app_data = _load_app_form(case, submission_folder)
    org = app_data.get("organization", {}) if app_data else {}
    project = app_data.get("project", {}) if app_data else {}
    funding = app_data.get("funding", {}) if app_data else {}

    fields = case.extracted_fields

    # Parse submission date
    try:
        sub_date = _parse_date(case.submission_timestamp) or date.today()
    except Exception:
        sub_date = date.today()

    # ---- R-002: Legal name cross-document mismatch ----------------------------
    _r002(findings, case, submission_folder, fields)

    # ---- R-003: Budget mismatch ------------------------------------------------
    _r003(case, submission_folder, findings, fields, funding, app_data)

    # ---- R-004: Project period -------------------------------------------------
    _r004(findings, fields)

    # ---- R-005: PacifiCan share > 75% -----------------------------------------
    _r005(findings, fields, case, submission_folder)

    # ---- R-006: Funding range --------------------------------------------------
    _r006(findings, fields)

    # ---- R-007: TRL for tech commercialization --------------------------------
    _r007(findings, fields, app_data, case, submission_folder)

    # ---- R-008: Tech Questionnaire required if tech_commercialization ---------
    _r008(findings, case, app_data)

    # ---- R-009: Low confidence fields -----------------------------------------
    _r009(findings, fields)

    # ---- R-010: Forecast/projection in funding confirmation -------------------
    _r010(findings, case, submission_folder)

    # ---- R-011: BC operating facilities = False --------------------------------
    _r011(findings, fields, org)

    # ---- R-012: Established < 2 years before submission -----------------------
    _r012(findings, org, sub_date)

    # ---- Eligibility flags ----------------------------------------------------
    flags = _build_eligibility_flags(
        app_data, org, project, funding, fields, case, submission_folder, sub_date
    )

    # De-duplicate R-002 (may have been added by extractor and rules)
    seen_r002 = False
    deduped: List[Finding] = []
    for f in findings:
        if f.rule_id == "R-002":
            if seen_r002:
                continue
            seen_r002 = True
        deduped.append(f)

    case.findings = deduped
    case.eligibility_flags = flags

    audit = AuditEvent(
        case_id=case.case_id,
        timestamp=_now_iso(),
        event_type=EventType.rules_evaluated,
        actor="system",
        details={
            "engine": "rules-legacy",
            "findings_count": len(deduped),
            "flags_count": len(flags),
            "rule_ids": list({f.rule_id for f in deduped}),
        },
    )
    case.audit_trail.append(audit)
    return case


# ===========================================================================
# MODEL / SKILL-BASED rules evaluation (default path when a key is configured)
# ===========================================================================

_RULES_SKILL_FILE = Path(__file__).parent / "rules.json"


def _load_rule_skills() -> Dict:
    try:
        with open(_RULES_SKILL_FILE, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def _serialise_fields(fields: Dict[str, ExtractedField]) -> List[Dict]:
    out: List[Dict] = []
    for fid, f in fields.items():
        out.append({
            "field_id": f.field_id,
            "name": f.name,
            "value": f.value,
            "confidence": f.confidence,
            "source_doc_id": f.source_doc_id,
            "evidence": f.raw_excerpt,
        })
    return out


def _run_rules_llm(case: Case, submission_folder: str) -> Case:
    """
    Model/skill-based rule evaluation guided by rules.json.

    One model call receives the extracted fields, the completeness checklist,
    the document texts, and the submission date, then applies every rule
    definition to produce findings + eligibility flags.  Raises on failure so
    the dispatcher can fall back to the dormant legacy engine.
    """
    from intake import llm
    from intake.extractor import _build_document_bundle

    rule_skills = _load_rule_skills()

    # Submission date
    try:
        sub_date = _parse_date(case.submission_timestamp) or date.today()
    except Exception:
        sub_date = date.today()

    # Completeness snapshot
    checklist = [
        {"doc_id": c.doc_id, "category": c.category, "status": c.status.value}
        for c in case.checklist
    ]
    tech_comm = any(
        c.doc_id == "DOC-08" and c.status != DocumentStatus.not_applicable
        for c in case.checklist
    ) if any(c.doc_id == "DOC-08" for c in case.checklist) else False

    # Document texts (capped) for content-dependent rules (R-002/003/007/010, ER-04)
    bundle = _build_document_bundle(case, submission_folder)
    docs_text = "\n\n".join(
        f"=== {b['doc_type']}: {b['label']} [{b['filename']}] ===\n{b['text'][:3000]}"
        for b in bundle
    )

    system_msg = (
        "You are a meticulous grant-eligibility officer. Apply every rule "
        "definition exactly as written, using the provided data. Do not invent "
        "rules. Return only the specified JSON object."
    )
    user_msg = (
        "RULE DEFINITIONS (apply each precisely; thresholds are authoritative):\n"
        f"{json.dumps(rule_skills, indent=2)}\n\n"
        f"SUBMISSION DATE: {sub_date.isoformat()}\n"
        f"TECHNOLOGY_COMMERCIALIZATION: {tech_comm}\n\n"
        "EXTRACTED FIELDS:\n"
        f"{json.dumps(_serialise_fields(case.extracted_fields), indent=2, default=str)}\n\n"
        "COMPLETENESS CHECKLIST:\n"
        f"{json.dumps(checklist, indent=2)}\n\n"
        "DOCUMENT TEXTS:\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{docs_text}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Return ONLY this JSON object:\n"
        "{\n"
        '  "findings": [ {"rule_id": <"R-0XX">, "severity": <"error"|"warning"|"info"|"manual_review">, "message": <str>} ],\n'
        '  "eligibility_flags": [ {"flag_id": <"ER-0X">, "label": <str>, "status": <"ok"|"needs_review"|"flagged">, "detail": <str>} ]\n'
        "}\n"
        "CRITICAL EMISSION RULE: Emit a finding ONLY when a rule's FAILURE or ATTENTION "
        "condition is triggered. Do NOT emit any finding when a rule passes cleanly — "
        "silence means the rule passed. "
        "The ONLY exception is R-002: always emit exactly one R-002 finding (warning / "
        "manual_review / info) to record the name-check outcome. "
        "All other rules (R-003 through R-012): zero findings if no issue detected. "
        "Produce all nine ER flags."
    )

    result = llm.chat_json(system_msg, user_msg, max_tokens=2500)
    if not result or "findings" not in result:
        raise RuntimeError("LLM rules evaluation returned no usable result")

    sev_map = {
        "error": Severity.error,
        "warning": Severity.warning,
        "info": Severity.info,
        "manual_review": Severity.manual_review,
    }
    findings: List[Finding] = []
    for item in result.get("findings", []):
        sev = sev_map.get(str(item.get("severity")), Severity.info)
        findings.append(Finding(
            rule_id=str(item.get("rule_id", "R-000")),
            severity=sev,
            message=str(item.get("message", "")),
        ))

    flags: List[EligibilityFlag] = []
    for item in result.get("eligibility_flags", []):
        flags.append(EligibilityFlag(
            flag_id=str(item.get("flag_id", "ER-00")),
            label=str(item.get("label", "")),
            status=str(item.get("status", "needs_review")),
            detail=str(item.get("detail", "")),
        ))

    case.findings = findings
    case.eligibility_flags = flags

    case.audit_trail.append(AuditEvent(
        case_id=case.case_id,
        timestamp=_now_iso(),
        event_type=EventType.rules_evaluated,
        actor="system",
        details={
            "engine": "model",
            "findings_count": len(findings),
            "flags_count": len(flags),
            "rule_ids": list({f.rule_id for f in findings}),
        },
    ))
    return case


def run_rules(case: Case, submission_folder: str) -> Case:
    """
    Dispatcher: model/skill-based evaluation when a key is configured,
    otherwise the (dormant) legacy deterministic engine as a safety net.
    """
    from intake import llm
    if llm.gpt_available():
        try:
            return _run_rules_llm(case, submission_folder)
        except Exception:
            return _run_rules_legacy(case, submission_folder)
    return _run_rules_legacy(case, submission_folder)


# ---------------------------------------------------------------------------
# Individual rule implementations
# ---------------------------------------------------------------------------

def _normalize_name(s: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace for name comparison."""
    import re as _re
    s = s.lower()
    s = _re.sub(r"[^\w\s]", "", s)
    return _re.sub(r"\s+", " ", s).strip()


def _r002(
    findings: List[Finding],
    case: Case,
    submission_folder: str,
    fields: Dict[str, ExtractedField],
) -> None:
    """R-002: Legal name cross-document mismatch.

    Three possible outcomes:
    1. Names compared and MATCH → no finding (passed)
    2. Names compared and DIFFER → warning finding
    3. Cross-document is PRESENT but name extraction failed → manual_review finding
    4. No cross-documents present at all → info finding (check skipped, not green)
    """
    folder = Path(submission_folder)
    app_name: Optional[str] = None
    supp_name: Optional[str] = None
    budget_name: Optional[str] = None

    # Get name from application form (DF-01 extracted value)
    df01 = fields.get("DF-01")
    if df01 and df01.value:
        app_name = str(df01.value)

    # Track which cross-documents are actually present in the submission
    supp_doc_present = False
    budget_doc_present = False

    # Get name from supplemental form (DOC-07) by reading PDF text
    for doc in case.documents:
        if doc.detected_doc_type == "DOC-07":
            supp_doc_present = True
            path = folder / doc.name
            if path.exists():
                text = _read_pdf_text(str(path))
                for line in text.splitlines():
                    if "LEGAL NAME" in line.upper():
                        parts = line.split(":", 1)
                        if len(parts) == 2:
                            supp_name = parts[1].strip()
                            break
            break

    # Get name from budget worksheet (DOC-04) — "Applicant:" header row
    for doc in case.documents:
        if doc.detected_doc_type == "DOC-04":
            budget_doc_present = True
            path = folder / doc.name
            if path.exists():
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(str(path), read_only=True, data_only=True)
                    for sheet_name in wb.sheetnames:
                        if sheet_name.lower() in ("instructions", "yearly"):
                            continue
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(max_row=10, values_only=True):
                            for i, cell in enumerate(row):
                                if cell and "applicant" in str(cell).lower():
                                    next_val = row[i + 1] if i + 1 < len(row) else None
                                    if next_val:
                                        budget_name = str(next_val).strip()
                                        break
                            if budget_name:
                                break
                        wb.close()
                except Exception:
                    pass
            break

    if not app_name:
        return

    # ── Case 3: doc present but name extraction failed ─────────────────────────
    # These are actionable: the officer must check manually.
    if supp_doc_present and not supp_name:
        findings.append(Finding(
            id="R-002-supp-extract-failed",
            rule_id="R-002",
            severity=Severity.manual_review,
            message=(
                "Legal name could not be parsed from supplemental form (DOC-07). "
                "Officer must manually verify the legal name matches the application form."
            ),
        ))
    if budget_doc_present and not budget_name:
        findings.append(Finding(
            id="R-002-budget-extract-failed",
            rule_id="R-002",
            severity=Severity.manual_review,
            message=(
                "Legal name could not be parsed from budget worksheet (DOC-04). "
                "Officer must manually verify the legal name matches the application form."
            ),
        ))

    # ── Case 4: no cross-documents at all ─────────────────────────────────────
    # This is a coverage gap, not a pass — flag it so it doesn't show green.
    if not supp_doc_present and not budget_doc_present:
        findings.append(Finding(
            id="R-002-no-cross-docs",
            rule_id="R-002",
            severity=Severity.info,
            message=(
                f"Name consistency check SKIPPED — neither the supplemental form (DOC-07) "
                f"nor the budget worksheet (DOC-04) was submitted. "
                f"Application form declares '{app_name}'. "
                f"Cross-document verification is not possible without these documents."
            ),
        ))
        return

    # ── Case 1 & 2: run the mismatch comparison ────────────────────────────────
    sources = {"application form": app_name}
    if supp_name:
        sources["supplemental form"] = supp_name
    if budget_name:
        sources["budget worksheet"] = budget_name

    norm_app = _normalize_name(app_name)
    for doc_label, name in sources.items():
        if doc_label == "application form":
            continue
        if _normalize_name(name) != norm_app:
            findings.append(Finding(
                id=f"R-002-{doc_label.replace(' ', '-')}",
                rule_id="R-002",
                severity=Severity.warning,
                message=(
                    f"Legal name in {doc_label} ('{name}') does not match "
                    f"application form ('{app_name}'). Officer review required."
                ),
            ))


def _r003(
    case: Case,
    submission_folder: str,
    findings: List[Finding],
    fields: Dict[str, ExtractedField],
    funding: Dict,
    app_data: Optional[Dict],
) -> None:
    """R-003: Budget worksheet total mismatch."""
    budget_path = _get_file_path(case, "DOC-04", submission_folder)
    if not budget_path:
        return

    current_total, application_total = _read_budget_totals(budget_path)
    if current_total is None or application_total is None:
        return

    diff = abs(current_total - application_total)
    if diff > 0.01:  # allow rounding
        findings.append(Finding(
            rule_id="R-003",
            severity=Severity.warning,
            message=(
                f"Budget worksheet Total Project Costs (Current) (${current_total:,.2f}) "
                f"does not match Total Project Costs (Application) (${application_total:,.2f}). "
                f"Difference: ${diff:,.2f}."
            ),
        ))

    # Also compare to application form total_project_costs
    if app_data:
        app_total = _unwrap(app_data.get("funding", {}).get("total_project_costs"))
        if app_total is not None:
            app_total_f = _parse_amount(app_total)
            budget_best = current_total or application_total
            if app_total_f is not None and budget_best is not None:
                cross_diff = abs(app_total_f - budget_best)
                if cross_diff > 0.01:
                    findings.append(Finding(
                        rule_id="R-003",
                        severity=Severity.warning,
                        message=(
                            f"Application form total project costs (${app_total_f:,.2f}) "
                            f"does not match budget worksheet (${budget_best:,.2f}). "
                            f"Difference: ${cross_diff:,.2f}."
                        ),
                    ))


def _r004(findings: List[Finding], fields: Dict[str, ExtractedField]) -> None:
    """R-004: Project period outside Apr 1 2026 – Mar 31 2028."""
    df07 = _get_field(fields, "DF-07")
    if not df07 or not df07.value:
        return

    # Handle both dict and string formats
    period = df07.value
    start_str = end_str = None

    if isinstance(period, dict):
        start_str = period.get("start_date")
        end_str = period.get("end_date")
    elif isinstance(period, str) and " to " in period:
        parts = period.split(" to ", 1)
        start_str = parts[0].strip()
        end_str = parts[1].strip()

    start = _parse_date(start_str)
    end = _parse_date(end_str)

    eligible_start = date(2026, 4, 1)
    eligible_end = date(2028, 3, 31)

    if start and start < eligible_start:
        findings.append(Finding(
            rule_id="R-004",
            severity=Severity.warning,
            message=(
                f"Project start date ({start.isoformat()}) is before the eligible "
                f"window start (April 1, 2026)."
            ),
        ))
    if end and end > eligible_end:
        findings.append(Finding(
            rule_id="R-004",
            severity=Severity.warning,
            message=(
                f"Project end date ({end.isoformat()}) is after the eligible "
                f"window end (March 31, 2028)."
            ),
        ))


def _r005(
    findings: List[Finding],
    fields: Dict[str, ExtractedField],
    case: Case,
    submission_folder: str,
) -> None:
    """R-005: PacifiCan share > 75%."""
    df05 = _get_field(fields, "DF-05")
    if not df05 or df05.value is None:
        return
    rda_requested = _parse_amount(df05.value)
    if rda_requested is None:
        return

    total_costs: Optional[float] = None
    budget_path = _get_file_path(case, "DOC-04", submission_folder)
    if budget_path:
        total_costs = _read_total_project_costs_from_budget(budget_path)

    if total_costs is None:
        df06 = _get_field(fields, "DF-06")
        if df06 and df06.value is not None:
            non_rda = _parse_amount(df06.value)
            if non_rda is not None:
                total_costs = rda_requested + non_rda

    if total_costs and total_costs > 0:
        share = rda_requested / total_costs
        if share > 0.75:
            findings.append(Finding(
                rule_id="R-005",
                severity=Severity.warning,
                message=(
                    f"PacifiCan funding share ({share:.1%}) exceeds the 75% maximum. "
                    f"Requested: ${rda_requested:,.2f}, Total: ${total_costs:,.2f}."
                ),
            ))


def _r006(findings: List[Finding], fields: Dict[str, ExtractedField]) -> None:
    """R-006: Funding outside $100K–$10M."""
    df05 = _get_field(fields, "DF-05")
    if not df05 or df05.value is None:
        return
    amount = _parse_amount(df05.value)
    if amount is None:
        return
    if amount < 100_000:
        findings.append(Finding(
            rule_id="R-006",
            severity=Severity.warning,
            message=f"Requested PacifiCan amount (${amount:,.2f}) is below the minimum of $100,000.",
        ))
    elif amount > 10_000_000:
        findings.append(Finding(
            rule_id="R-006",
            severity=Severity.warning,
            message=f"Requested PacifiCan amount (${amount:,.2f}) exceeds the maximum of $10,000,000.",
        ))


def _r007(
    findings: List[Finding],
    fields: Dict[str, ExtractedField],
    app_data: Optional[Dict],
    case: Case,
    submission_folder: str,
) -> None:
    """R-007: TRL < 5 for tech commercialization."""
    if not app_data:
        return
    project = app_data.get("project", {})
    tech_comm = bool(_unwrap(project.get("technology_commercialization", False)))
    if not tech_comm:
        return

    trl: Optional[int] = None
    raw_trl = project.get("trl") or project.get("technology_readiness_level")
    if raw_trl is not None:
        try:
            trl = int(_unwrap(raw_trl))
        except (TypeError, ValueError):
            pass

    if trl is None:
        tech_path = _get_file_path(case, "DOC-08", submission_folder)
        if tech_path:
            text = _read_pdf_text(tech_path)
            m = re.search(r"TRL[:\s]+(\d+)", text, re.IGNORECASE)
            if m:
                try:
                    trl = int(m.group(1))
                except ValueError:
                    pass

    if trl is None:
        findings.append(Finding(
            rule_id="R-007",
            severity=Severity.warning,
            message="TRL not declared for Technology Commercialization project.",
        ))
    elif trl < 5:
        findings.append(Finding(
            rule_id="R-007",
            severity=Severity.warning,
            message=f"TRL {trl} is below the minimum threshold of TRL 5 for Technology Commercialization projects.",
        ))


def _r008(
    findings: List[Finding],
    case: Case,
    app_data: Optional[Dict],
) -> None:
    """R-008: Tech Questionnaire required if technology_commercialization=True."""
    if not app_data:
        return
    project = app_data.get("project", {})
    tech_comm = bool(_unwrap(project.get("technology_commercialization", False)))
    if not tech_comm:
        return

    has_doc08 = any(
        d.detected_doc_type == "DOC-08" and d.confidence >= 0.70
        for d in case.documents
    )
    if not has_doc08:
        findings.append(Finding(
            rule_id="R-008",
            severity=Severity.error,
            message=(
                "RDII Technology Questionnaire (DOC-08) is required for Technology "
                "Commercialization projects but is missing from the submission."
            ),
        ))


def _r009(findings: List[Finding], fields: Dict[str, ExtractedField]) -> None:
    """R-009: Confidence < 0.60 or value None → MANUAL_REVIEW per field."""
    for field_id, field in fields.items():
        # Flag if confidence is low OR if value is missing (null) and confidence is 0
        needs_review = field.confidence < 0.60
        if needs_review:
            if field.value is None:
                msg = (
                    f"Field {field.field_id} ({field.name}) could not be extracted "
                    f"(value not found). Manual entry required."
                )
            else:
                msg = (
                    f"Field {field.field_id} ({field.name}) has low extraction "
                    f"confidence ({field.confidence:.0%}). Manual review required."
                )
            findings.append(Finding(
                rule_id="R-009",
                severity=Severity.manual_review,
                message=msg,
            ))


def _r010(findings: List[Finding], case: Case, submission_folder: str) -> None:
    """R-010: Forecast/projection language in funding confirmation."""
    doc06_path = _get_file_path(case, "DOC-06", submission_folder)
    if not doc06_path:
        return

    text = _read_pdf_text(doc06_path).lower()
    if not text:
        return

    forecast_terms = ["forecast", "projection", "anticipated", "expected revenue", "pro forma"]
    confirmed_terms = ["awarded", "confirmed", "signed", "letter of agreement"]

    has_forecast = any(term in text for term in forecast_terms)
    has_confirmed = any(term in text for term in confirmed_terms)

    if has_forecast and not has_confirmed:
        found = [t for t in forecast_terms if t in text]
        findings.append(Finding(
            rule_id="R-010",
            severity=Severity.warning,
            message=(
                f"Funding confirmation (DOC-06) contains forecast/projection language "
                f"({', '.join(found)}) without confirmed-funding language. "
                f"Officer review required."
            ),
        ))


def _r011(
    findings: List[Finding],
    fields: Dict[str, ExtractedField],
    org: Dict,
) -> None:
    """R-011: bc_operating_facilities = False → INFO."""
    df04 = _get_field(fields, "DF-04")
    bc_present: Optional[bool] = None

    if df04 and df04.value and isinstance(df04.value, dict):
        raw_bc = df04.value.get("bc_operating_facilities")
        if raw_bc is not None:
            bc_present = bool(raw_bc)

    if bc_present is None:
        raw = _unwrap(org.get("bc_operating_facilities"))
        if raw is not None:
            bc_present = bool(raw)

    if bc_present is False:
        findings.append(Finding(
            rule_id="R-011",
            severity=Severity.info,
            message="Applicant does not declare a BC operating facility. Confirm BC nexus and eligibility.",
        ))


def _r012(findings: List[Finding], org: Dict, submission_date: date) -> None:
    """R-012: Established in Canada + 2 years > submission_date → INFO."""
    raw = _unwrap(
        org.get("date_established_in_canada")
        or org.get("incorporation_date")
        or org.get("date_established")
    )
    established = _parse_date(raw)
    if established is None:
        return

    try:
        from dateutil.relativedelta import relativedelta
        two_years_after = established + relativedelta(years=2)
    except Exception:
        import datetime as _dt
        two_years_after = established.replace(year=established.year + 2)

    if two_years_after > submission_date:
        findings.append(Finding(
            rule_id="R-012",
            severity=Severity.info,
            message=(
                f"Organisation established {established.isoformat()} — less than 2 years "
                f"before submission ({submission_date.isoformat()}). Operating history requirement "
                f"may not be met."
            ),
        ))


# ---------------------------------------------------------------------------
# Eligibility flags
# ---------------------------------------------------------------------------

def _build_eligibility_flags(
    app_data: Optional[Dict],
    org: Dict,
    project: Dict,
    funding: Dict,
    fields: Dict[str, ExtractedField],
    case: Case,
    submission_folder: str,
    submission_date: date,
) -> List[EligibilityFlag]:
    flags: List[EligibilityFlag] = []

    # ER-01: Recipient type
    corporate_status = str(_unwrap(org.get("corporate_status") or org.get("organization_type", "")))
    if corporate_status:
        lower = corporate_status.lower()
        if any(w in lower for w in ("for-profit", "for profit", "corporation", "inc", "ltd")):
            flags.append(EligibilityFlag(flag_id="ER-01", label="Recipient Type", status="ok",
                                          detail=f"Corporate status: {corporate_status}"))
        else:
            flags.append(EligibilityFlag(flag_id="ER-01", label="Recipient Type", status="needs_review",
                                          detail=f"Corporate status: {corporate_status}"))
    else:
        flags.append(EligibilityFlag(flag_id="ER-01", label="Recipient Type", status="needs_review",
                                      detail="Corporate status not found."))

    # ER-02: Operating history
    raw_est = _unwrap(
        org.get("date_established_in_canada")
        or org.get("incorporation_date")
        or org.get("date_established")
    )
    established = _parse_date(raw_est)
    if established:
        try:
            from dateutil.relativedelta import relativedelta
            two_years = established + relativedelta(years=2)
        except Exception:
            two_years = established.replace(year=established.year + 2)
        if two_years <= submission_date:
            flags.append(EligibilityFlag(flag_id="ER-02", label="Operating History (2+ years)", status="ok",
                                          detail=f"Established {established.isoformat()}."))
        else:
            flags.append(EligibilityFlag(flag_id="ER-02", label="Operating History (2+ years)", status="flagged",
                                          detail=f"Established {established.isoformat()} — less than 2 years before submission."))
    else:
        flags.append(EligibilityFlag(flag_id="ER-02", label="Operating History (2+ years)", status="needs_review",
                                      detail="Establishment date not found."))

    # ER-03: BC presence
    raw_bc = _unwrap(org.get("bc_operating_facilities"))
    df04 = fields.get("DF-04")
    if df04 and isinstance(df04.value, dict):
        raw_bc = df04.value.get("bc_operating_facilities", raw_bc)
    if raw_bc is True:
        flags.append(EligibilityFlag(flag_id="ER-03", label="BC Presence", status="ok",
                                      detail="BC operating facility declared."))
    elif raw_bc is False:
        flags.append(EligibilityFlag(flag_id="ER-03", label="BC Presence", status="flagged",
                                      detail="No BC operating facility declared."))
    else:
        flags.append(EligibilityFlag(flag_id="ER-03", label="BC Presence", status="needs_review",
                                      detail="BC facility status not found."))

    # ER-04: Defence supply chain
    desc_text = str(_unwrap(project.get("description") or project.get("project_description") or "")).lower()
    supp_path = _get_file_path(case, "DOC-07", submission_folder)
    if supp_path:
        desc_text += " " + _read_pdf_text(supp_path).lower()
    defence_kw = ["defence", "defense", "military", "dnd", "nato", "supply chain"]
    if any(kw in desc_text for kw in defence_kw):
        flags.append(EligibilityFlag(flag_id="ER-04", label="Defence Supply Chain", status="ok",
                                      detail="Defence supply chain keywords found."))
    else:
        flags.append(EligibilityFlag(flag_id="ER-04", label="Defence Supply Chain", status="needs_review",
                                      detail="Defence supply chain connection not evident. Manual review required."))

    # ER-05: Defence capability (always manual)
    flags.append(EligibilityFlag(flag_id="ER-05", label="Defence Capability", status="needs_review",
                                  detail="Officer must assess defence capability alignment."))

    # ER-06: Project period
    df07 = fields.get("DF-07")
    er06_status = "ok"
    er06_detail = "Project period within eligible window (Apr 1, 2026 – Mar 31, 2028)."
    if df07 and df07.value:
        period = df07.value
        start_str = end_str = None
        if isinstance(period, dict):
            start_str = period.get("start_date")
            end_str = period.get("end_date")
        elif isinstance(period, str) and " to " in period:
            parts = period.split(" to ", 1)
            start_str, end_str = parts[0].strip(), parts[1].strip()
        s = _parse_date(start_str)
        e = _parse_date(end_str)
        issues = []
        if s and s < date(2026, 4, 1):
            issues.append(f"Start {s.isoformat()} before Apr 1, 2026")
        if e and e > date(2028, 3, 31):
            issues.append(f"End {e.isoformat()} after Mar 31, 2028")
        if issues:
            er06_status = "flagged"
            er06_detail = "; ".join(issues)
    flags.append(EligibilityFlag(flag_id="ER-06", label="Project Period", status=er06_status, detail=er06_detail))

    # ER-07: TRL
    tech_comm = bool(_unwrap(project.get("technology_commercialization", False)))
    if not tech_comm:
        flags.append(EligibilityFlag(flag_id="ER-07", label="Technology Readiness Level", status="ok",
                                      detail="Not applicable (not a Tech Commercialization project)."))
    else:
        raw_trl = _unwrap(project.get("trl") or project.get("technology_readiness_level"))
        trl = None
        if raw_trl is not None:
            try:
                trl = int(raw_trl)
            except (TypeError, ValueError):
                pass
        if trl is None:
            flags.append(EligibilityFlag(flag_id="ER-07", label="TRL", status="needs_review", detail="TRL not declared."))
        elif trl < 5:
            flags.append(EligibilityFlag(flag_id="ER-07", label="TRL", status="flagged", detail=f"TRL {trl} below minimum of 5."))
        else:
            flags.append(EligibilityFlag(flag_id="ER-07", label="TRL", status="ok", detail=f"TRL {trl} meets minimum."))

    # ER-08: Funding range
    df05 = fields.get("DF-05")
    er08_status = "ok"
    er08_detail = "Requested amount within eligible range ($100K–$10M)."
    if df05 and df05.value is not None:
        amt = _parse_amount(df05.value)
        if amt is not None and (amt < 100_000 or amt > 10_000_000):
            er08_status = "flagged"
            er08_detail = f"${amt:,.2f} is outside eligible range."
    flags.append(EligibilityFlag(flag_id="ER-08", label="Funding Range", status=er08_status, detail=er08_detail))

    # ER-09: PacifiCan share ≤ 75%
    er09_status = "ok"
    er09_detail = "PacifiCan share within 75% limit."
    df06 = fields.get("DF-06")
    if df05 and df05.value is not None and df06 and df06.value is not None:
        rda = _parse_amount(df05.value)
        non_rda = _parse_amount(df06.value)
        if rda is not None and non_rda is not None:
            total = rda + non_rda
            if total > 0:
                share = rda / total
                if share > 0.75:
                    er09_status = "flagged"
                    er09_detail = f"PacifiCan share {share:.1%} exceeds 75% limit."
                else:
                    er09_detail = f"PacifiCan share {share:.1%} is within limit."
    flags.append(EligibilityFlag(flag_id="ER-09", label="PacifiCan Share ≤ 75%", status=er09_status, detail=er09_detail))

    return flags
